#!/usr/bin/env python3
"""
URL 우선순위 목록 생성기

koreaners.co의 모든 페이지 URL을 추출하고 우선순위를 지정하여
CSV 및 Excel 파일로 출력합니다.

사용법:
    python url_priority_generator.py

출력 파일:
    - url_priority_list.csv
    - indexing_schedule.xlsx (openpyxl 설치 시)
"""

import csv
import os
import sys
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import json

# Supabase 클라이언트 (필요시)
try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False
    print("⚠️  supabase-py 미설치. pip install supabase 실행")

# Excel 지원 (선택적)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl 미설치. pip install openpyxl 실행 (Excel 출력 필요시)")


# 환경 변수
SUPABASE_URL = os.getenv('NEXT_PUBLIC_SUPABASE_URL', '')
SUPABASE_KEY = os.getenv('NEXT_PUBLIC_SUPABASE_ANON_KEY', '')
BASE_URL = 'https://www.koreaners.co'


class URLPriorityGenerator:
    """URL 우선순위 생성 및 관리 클래스"""

    def __init__(self):
        self.base_url = BASE_URL
        self.urls: List[Dict] = []
        self.supabase: Optional[Client] = None

        if SUPABASE_AVAILABLE and SUPABASE_URL and SUPABASE_KEY:
            self.supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    def add_static_pages(self):
        """정적 페이지 추가"""
        static_pages = [
            {
                'url': self.base_url,
                'title': '코리너스 홈',
                'priority': 'Critical',
                'priority_score': 1.0,
                'estimated_traffic': '높음',
                'category': '메인',
                'indexing_status': '확인 필요',
                'index_date': None,
            },
            {
                'url': f'{self.base_url}/portfolio',
                'title': '포트폴리오',
                'priority': 'Critical',
                'priority_score': 0.9,
                'estimated_traffic': '높음',
                'category': '주요 서비스',
                'indexing_status': '확인 필요',
                'index_date': None,
            },
            {
                'url': f'{self.base_url}/blog',
                'title': '블로그',
                'priority': 'Critical',
                'priority_score': 0.9,
                'estimated_traffic': '높음',
                'category': '주요 서비스',
                'indexing_status': '확인 필요',
                'index_date': None,
            },
            {
                'url': f'{self.base_url}/creator',
                'title': '크리에이터 합류',
                'priority': 'High',
                'priority_score': 0.8,
                'estimated_traffic': '중간',
                'category': '서비스',
                'indexing_status': '확인 필요',
                'index_date': None,
            },
            {
                'url': f'{self.base_url}/inquiry',
                'title': '문의하기',
                'priority': 'Medium',
                'priority_score': 0.7,
                'estimated_traffic': '중간',
                'category': '서비스',
                'indexing_status': '확인 필요',
                'index_date': None,
            },
        ]

        self.urls.extend(static_pages)
        print(f"✅ 정적 페이지 {len(static_pages)}개 추가")

    def add_dynamic_pages(self):
        """동적 페이지 추가 (Supabase에서 가져오기)"""
        if not self.supabase:
            print("⚠️  Supabase 연결 없음. 동적 페이지 스킵")
            return

        try:
            # 포트폴리오 페이지
            portfolios = self.supabase.table('portfolios').select('id, title, created_at').execute()
            for portfolio in portfolios.data:
                self.urls.append({
                    'url': f'{self.base_url}/portfolio/{portfolio["id"]}',
                    'title': f'포트폴리오: {portfolio.get("title", "제목 없음")}',
                    'priority': 'High',
                    'priority_score': 0.8,
                    'estimated_traffic': '중간',
                    'category': '포트폴리오 상세',
                    'indexing_status': '확인 필요',
                    'index_date': None,
                })
            print(f"✅ 포트폴리오 페이지 {len(portfolios.data)}개 추가")

            # 블로그 포스트
            blog_posts = self.supabase.table('blog_posts').select('id, title, published, created_at').eq('published', True).execute()
            for post in blog_posts.data:
                self.urls.append({
                    'url': f'{self.base_url}/blog/{post["id"]}',
                    'title': f'블로그: {post.get("title", "제목 없음")}',
                    'priority': 'Medium',
                    'priority_score': 0.7,
                    'estimated_traffic': '낮음-중간',
                    'category': '블로그 포스트',
                    'indexing_status': '확인 필요',
                    'index_date': None,
                })
            print(f"✅ 블로그 포스트 {len(blog_posts.data)}개 추가")

            # 크리에이터 프로필
            creators = self.supabase.table('creators').select('id, name, created_at').execute()
            for creator in creators.data:
                self.urls.append({
                    'url': f'{self.base_url}/creator/{creator["id"]}',
                    'title': f'크리에이터: {creator.get("name", "이름 없음")}',
                    'priority': 'Low',
                    'priority_score': 0.6,
                    'estimated_traffic': '낮음',
                    'category': '크리에이터 프로필',
                    'indexing_status': '확인 필요',
                    'index_date': None,
                })
            print(f"✅ 크리에이터 프로필 {len(creators.data)}개 추가")

        except Exception as e:
            print(f"❌ 동적 페이지 가져오기 실패: {e}")

    def calculate_index_dates(self):
        """우선순위에 따른 인덱싱 예상 날짜 계산"""
        priority_schedule = {
            'Critical': 0,  # 즉시
            'High': 7,      # 1주일 이내
            'Medium': 14,   # 2주일 이내
            'Low': 30,      # 1개월 이내
        }

        today = datetime.now()
        for url_info in self.urls:
            priority = url_info['priority']
            days_offset = priority_schedule.get(priority, 30)
            url_info['index_date'] = (today + timedelta(days=days_offset)).strftime('%Y-%m-%d')

    def sort_by_priority(self):
        """우선순위 순으로 정렬"""
        self.urls.sort(key=lambda x: x['priority_score'], reverse=True)

    def export_to_csv(self, filename='url_priority_list.csv'):
        """CSV 파일로 출력"""
        if not self.urls:
            print("❌ 출력할 URL이 없습니다")
            return

        fieldnames = [
            'url', 'title', 'priority', 'priority_score',
            'estimated_traffic', 'category', 'indexing_status', 'index_date'
        ]

        with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.urls)

        print(f"✅ CSV 파일 생성: {filename} ({len(self.urls)}개 URL)")

    def export_to_excel(self, filename='indexing_schedule.xlsx'):
        """Excel 파일로 출력 (고급 포맷팅)"""
        if not EXCEL_AVAILABLE:
            print("⚠️  Excel 출력 불가: openpyxl 미설치")
            return

        if not self.urls:
            print("❌ 출력할 URL이 없습니다")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "인덱싱 스케줄"

        # 헤더 스타일
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # 헤더 작성
        headers = ['URL', '페이지 제목', '우선순위', '우선순위 점수',
                   '예상 트래픽', '카테고리', '인덱싱 상태', '예상 인덱싱 날짜']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 데이터 작성
        for row_num, url_info in enumerate(self.urls, 2):
            ws.cell(row=row_num, column=1, value=url_info['url'])
            ws.cell(row=row_num, column=2, value=url_info['title'])
            ws.cell(row=row_num, column=3, value=url_info['priority'])
            ws.cell(row=row_num, column=4, value=url_info['priority_score'])
            ws.cell(row=row_num, column=5, value=url_info['estimated_traffic'])
            ws.cell(row=row_num, column=6, value=url_info['category'])
            ws.cell(row=row_num, column=7, value=url_info['indexing_status'])
            ws.cell(row=row_num, column=8, value=url_info['index_date'])

            # 우선순위별 색상
            priority_colors = {
                'Critical': 'FF6B6B',
                'High': 'FFA500',
                'Medium': 'FFD93D',
                'Low': '95E1D3',
            }
            priority_fill = PatternFill(
                start_color=priority_colors.get(url_info['priority'], 'FFFFFF'),
                end_color=priority_colors.get(url_info['priority'], 'FFFFFF'),
                fill_type='solid'
            )
            ws.cell(row=row_num, column=3).fill = priority_fill

        # 열 너비 자동 조정
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(headers[col_num - 1])
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(filename)
        print(f"✅ Excel 파일 생성: {filename} ({len(self.urls)}개 URL)")

    def generate(self):
        """전체 프로세스 실행"""
        print("🚀 URL 우선순위 목록 생성 시작\n")

        self.add_static_pages()
        self.add_dynamic_pages()
        self.calculate_index_dates()
        self.sort_by_priority()

        print(f"\n📊 총 {len(self.urls)}개 URL 생성 완료\n")

        # 우선순위별 통계
        priority_counts = {}
        for url_info in self.urls:
            priority = url_info['priority']
            priority_counts[priority] = priority_counts.get(priority, 0) + 1

        print("📈 우선순위별 통계:")
        for priority, count in sorted(priority_counts.items(),
                                     key=lambda x: ['Critical', 'High', 'Medium', 'Low'].index(x[0])):
            print(f"   - {priority}: {count}개")

        print()
        self.export_to_csv()
        self.export_to_excel()

        print("\n✅ 완료!")


def main():
    """메인 실행 함수"""
    generator = URLPriorityGenerator()
    generator.generate()


if __name__ == '__main__':
    main()
